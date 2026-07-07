#!/usr/bin/env python3
"""
Main HUB → DHL 用车记录 字段匹配工具
GUI - tkinter + openpyxl
支持图片复制、跨Sheet重复Reference检测
打包: pyinstaller --onefile --noconsole --name "DHL匹配工具" match_tool.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, json, os, re, shutil
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker

# ============================================================ 工具函数 ============================================================

def norm(text):
    if text is None: return ""
    return " ".join(str(text).strip().split())

def fuzzy_find(headers, hint):
    hlow = hint.lower()
    for i, h in enumerate(headers):
        n = norm(h).lower()
        if not n or n == "none": continue
        if n == hlow: return i
        if hlow in n or n in hlow: return i
    return -1

def read_sheet_rows(wb, sheet_name):
    ws = wb[sheet_name]
    data = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if not data: return [], []
    headers = [str(c) if c is not None else "" for c in data[0]]
    rows = []
    for r in data[1:]:
        if r is None: continue
        row = [c if c is not None else "" for c in r]
        if all(str(v).strip() == "" for v in row): continue
        rows.append(row)
    return headers, rows

def read_sheet_headers(wb, sheet_name):
    """只读表头，不加载数据行"""
    ws = wb[sheet_name]
    data = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not data: return []
    return [str(c) if c is not None else "" for c in data[0]]

def _find_candidate_sheets(wb):
    candidates = []
    for name in wb.sheetnames:
        try:
            headers = read_sheet_headers(wb, name)
            if not headers: continue
            has_lr = any("loading reference" in norm(h).lower() or "提货码" in norm(h).lower() for h in headers)
            if has_lr and len(headers) >= 20:
                candidates.append(name)
        except Exception: pass
    return candidates

def is_valid_reference(val):
    v = str(val).strip()
    if not v: return False
    placeholders = {"-", "/", "none", "cancel", "need cancel", "empty", "n/a", "na", "tbd", "x", ""}
    if v.lower() in placeholders: return False
    if "TEM" not in v.upper(): return False
    return True

def is_strikethrough(cell):
    """判断 openpyxl Cell 是否有删除线（字体 strikethrough 属性）"""
    return cell.font and cell.font.strike

# ============================================================ 数据加载 & 重复检测 ============================================================

def read_source_data(wb, selected_sheets):
    """返回 (表头, 总行数, 图片映射)。只读表头和计数，不将所有数据行加载到内存。"""
    first_headers = None
    total_rows = 0
    image_map = {}
    for sn in selected_sheets:
        ws = wb[sn]
        headers = read_sheet_headers(wb, sn)
        if first_headers is None: first_headers = headers
        # 统计有效行数（跳过空行和表头行）
        for row in ws.iter_rows(min_row=2, values_only=True):
            row = [c if c is not None else "" for c in row]
            if not all(str(v).strip() == "" for v in row):
                total_rows += 1
        if hasattr(ws, "_images") and ws._images:
            for img in ws._images:
                try:
                    fc = img.anchor._from
                    # _from.row 是 0-based sheet 行索引（含表头），减 1 后与
                    # iter_source_rows 的 ri（0-based 数据行索引）对齐
                    image_map.setdefault((sn, fc.row - 1), []).append((img, fc.col))
                except Exception: pass
    return first_headers, total_rows, image_map

def iter_source_rows(wb, selected_sheets):
    """流式迭代源数据行，返回 (sheet_name, row_0based_index, row_data)"""
    for sn in selected_sheets:
        ws = wb[sn]
        for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            row = [c if c is not None else "" for c in row]
            if all(str(v).strip() == "" for v in row): continue
            yield (sn, ri, row)

def check_duplicates(selected_sheets, wb):
    if len(selected_sheets) <= 1: return {}
    all_refs = defaultdict(list)
    for sn in selected_sheets:
        ws = wb[sn]
        headers = read_sheet_headers(wb, sn)
        if not headers: continue
        li = fuzzy_find(headers, "Loading Reference")
        if li == -1: li = fuzzy_find(headers, "提货码")
        if li == -1: continue
        for ri, row in enumerate(ws.iter_rows(min_row=2)):
            cell = row[li] if li < len(row) else None
            if cell is None or cell.value is None: continue
            if is_strikethrough(cell): continue
            ref = str(cell.value).strip()
            if is_valid_reference(ref):
                all_refs[ref].append((sn, ri + 2))
    result = {}
    for ref, locs in all_refs.items():
        sheets_ = set(s[0] for s in locs)
        if len(sheets_) >= 2:
            result[ref] = locs
    return result

# ============================================================ 图片复制 ============================================================

def copy_images_for_row(source_wb, sheet_name, src_row_0based, target_ws, tgt_row_1based,
                        src_col_idx, tgt_col_idx, image_map):
    key = (sheet_name, src_row_0based)
    if key not in image_map: return 0
    copied = 0
    for img, img_col in image_map[key]:
        if img_col != src_col_idx: continue
        try:
            # 确保每次读取时流位置正确（Windows 上 openpyxl 的 ref 可能未重置）
            img_data = img.ref
            if hasattr(img_data, 'seek'):
                img_data.seek(0)
            new_img = XLImage(img_data)
            new_img.width = float(img.width) if img.width else 100
            new_img.height = float(img.height) if img.height else 100
            marker = AnchorMarker(col=tgt_col_idx, colOff=0, row=tgt_row_1based - 1, rowOff=0)
            ext = img.anchor.ext if hasattr(img.anchor, 'ext') else None
            new_img.anchor = OneCellAnchor(_from=marker, ext=ext)
            target_ws.add_image(new_img)
            copied += 1
        except Exception: pass
    return copied

# ============================================================ 默认映射 ============================================================

DEFAULT_MAPPINGS = [
    ("Destination", "Destination"),
    ("AWB", "Bill of Lading No. (AWB, B/L)"),
    ("Plate/车牌", "License Plate No."),
    ("Actual Date & Time \nfor arrival", "ATA (Actual Arrival)"),
    ("Acual Date & Time \nfor arrival", "ATA (Actual Arrival)"),  # 兼容源数据中可能存在的拼写错误
    ("Planned Date & Time \nfor loading", "ETA (Planned Pickup)"),
    ("Planned Time \nfor loading", "ETA (Planned Pickup)"),
    ("Cartons", "Number of Boxes"),
    ("Note", "Note"),
    ("Weight", "Weight"),
    ("Price/车费", "Price"),
    ("Trucker/车队", "Trucker"),
]

# DP代码 → 目的地城市（可从 Reference 中 DP 后面的代码解析真实目的地）
DP_TO_CITY = {
    "BOCH": "Bochum",
    "BRUC": "Bruchsal",
    "DORS": "Dorsten",
    "EUTI": "Eutingen",
    "GREV": "Greven",
    "HAGE": "Hagen",
    "HANN": "Hannover",
    "KITZ": "Kitzingen",
    "KÖNG": "Köngen",
    "LAHR": "Lahr",
    "NEUW": "Neuwied",
    "OBER": "Obertshausen",
    "SAUL": "Saulheim",
    "SPEY": "Speyer",
    "STAU": "Staufenberg",
}

def extract_destination(ref, dp_map=None):
    """从 Reference 中提取 DP 代码并映射为目的地城市名。
    例: TEM-SKY-DPHAGE-060726-4 → Hagen
    如果 dp_map 为 None，使用全局 DP_TO_CITY"""
    if dp_map is None:
        dp_map = DP_TO_CITY
    m = re.search(r'DP(\w+?)-', str(ref))
    if m:
        code = m.group(1)
        if code in dp_map:
            return dp_map[code]
        else:
            return f"??{code}"  # 未知代码，标记以便人工确认
    return ""

def _dp_config_path():
    """DP 映射配置文件的路径，保存到用户目录下（避免 exe 所在目录无写入权限）"""
    import sys
    if sys.platform == "win32":
        base = os.environ.get("APPDATA", os.path.expanduser("~"))
        dir_name = "DHL-MatchTool"
    else:
        base = os.path.expanduser("~")
        dir_name = ".DHL-MatchTool"
    cfg_dir = os.path.join(base, dir_name)
    os.makedirs(cfg_dir, exist_ok=True)
    return os.path.join(cfg_dir, "dp_config.json")

def load_dp_map():
    """从配置文件加载 DP 映射，如果文件不存在或损坏则使用默认值"""
    path = _dp_config_path()
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict) and data:
                return data
    except (FileNotFoundError, json.JSONDecodeError):
        pass
    return dict(DP_TO_CITY)

def save_dp_map(dp_map):
    """将 DP 映射保存到配置文件"""
    try:
        with open(_dp_config_path(), "w", encoding="utf-8") as f:
            json.dump(dp_map, f, ensure_ascii=False, indent=2)
    except Exception:
        pass  # 保存失败不阻塞

# ============================================================ GUI ============================================================

class MatchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Main HUB → DHL 用车记录 字段匹配工具")
        self.root.geometry("1000x800")
        self.root.minsize(850, 650)
        self.style = ttk.Style()
        self.style.theme_use("clam")

        self.src_wb = None; self.src_path = None
        self.src_headers = None; self.src_row_count = 0; self.src_images = None
        self.selected_sheets = ["LEJ", "LGG"]
        self.tgt_wb = None; self.tgt_path = None
        self.tgt_headers = None; self.tgt_rows = None; self.tgt_sheet = None
        self.mapping_rows = []
        self.dp_map = load_dp_map()          # 从文件加载 DP 映射（持久化）
        self._has_cross_sheet_dups = False  # 是否有跨Sheet重复

        self._build_ui()

    # ==================== UI 构建 ====================
    def _build_ui(self):
        main = ttk.Frame(self.root, padding=12)
        main.pack(fill=tk.BOTH, expand=True)

        # ── 文件选择 ──
        ff = ttk.LabelFrame(main, text="📂 文件选择", padding=10)
        ff.pack(fill=tk.X, pady=(0, 6))

        sf = ttk.Frame(ff); sf.pack(fill=tk.X, pady=2)
        ttk.Label(sf, text="源文件 (Main HUB):", width=18).pack(side=tk.LEFT)
        self.src_path_var = tk.StringVar()
        ttk.Entry(sf, textvariable=self.src_path_var, width=50).pack(side=tk.LEFT, padx=4)
        ttk.Button(sf, text="选择...", command=self._select_src, width=8).pack(side=tk.LEFT)
        self.src_st = tk.StringVar(value="未加载")
        ttk.Label(sf, textvariable=self.src_st, foreground="gray", width=22).pack(side=tk.LEFT, padx=8)

        tf = ttk.Frame(ff); tf.pack(fill=tk.X, pady=2)
        ttk.Label(tf, text="目标文件 (DHL 记录):", width=18).pack(side=tk.LEFT)
        self.tgt_path_var = tk.StringVar()
        ttk.Entry(tf, textvariable=self.tgt_path_var, width=50).pack(side=tk.LEFT, padx=4)
        ttk.Button(tf, text="选择...", command=self._select_tgt, width=8).pack(side=tk.LEFT)
        self.tgt_st = tk.StringVar(value="未加载")
        ttk.Label(tf, textvariable=self.tgt_st, foreground="gray", width=22).pack(side=tk.LEFT, padx=8)

        # ── Sheet 选择 ──
        self.shf = ttk.LabelFrame(main, text="📑 源数据 Sheet – 勾选要合并的页面（数据自动合并）", padding=10)
        self.shf.pack(fill=tk.X, pady=(0, 6))
        self.shcf = ttk.Frame(self.shf); self.shcf.pack(fill=tk.X)

        # ── 匹配配置 ──
        cf = ttk.LabelFrame(main, text="🔗 匹配配置", padding=10)
        cf.pack(fill=tk.X, pady=(0, 6))

        # 匹配键
        kf = ttk.Frame(cf); kf.pack(fill=tk.X, pady=2)
        ttk.Label(kf, text="源匹配键:").pack(side=tk.LEFT)
        self.src_key_var = tk.StringVar()
        self.src_key_cb = ttk.Combobox(kf, textvariable=self.src_key_var, width=28, state="readonly")
        self.src_key_cb.pack(side=tk.LEFT, padx=4)
        ttk.Label(kf, text="⬌").pack(side=tk.LEFT, padx=8)
        ttk.Label(kf, text="目标匹配键:").pack(side=tk.LEFT)
        self.tgt_key_var = tk.StringVar()
        self.tgt_key_cb = ttk.Combobox(kf, textvariable=self.tgt_key_var, width=28, state="readonly")
        self.tgt_key_cb.pack(side=tk.LEFT, padx=4)

        # 映射表头
        mh = ttk.Frame(cf); mh.pack(fill=tk.X, pady=(10, 2))
        ttk.Label(mh, text="#", width=3).pack(side=tk.LEFT)
        ttk.Label(mh, text="源字段 (Main HUB)", width=38).pack(side=tk.LEFT, padx=4)
        ttk.Label(mh, text="→", width=3).pack(side=tk.LEFT)
        ttk.Label(mh, text="目标字段 (DHL)", width=38).pack(side=tk.LEFT, padx=4)

        # ==== 可滚动的映射区域（关键修复：带滚动条）====
        map_container = ttk.Frame(cf)
        map_container.pack(fill=tk.BOTH, expand=True, pady=2)

        self.map_canvas = tk.Canvas(map_container, height=140, highlightthickness=0, bg="#2a2a3e")
        self.map_scrollbar = ttk.Scrollbar(map_container, orient=tk.VERTICAL, command=self.map_canvas.yview)
        self.map_canvas.configure(yscrollcommand=self.map_scrollbar.set)
        self.map_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.map_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.map_inner = ttk.Frame(self.map_canvas)
        self.map_canvas_window = self.map_canvas.create_window((0, 0), window=self.map_inner, anchor="nw")
        self.map_inner.bind("<Configure>", self._on_map_inner_configure)
        self.map_canvas.bind("<Configure>", self._on_map_canvas_configure)

        # 按钮行 + 状态
        bf = ttk.Frame(cf); bf.pack(fill=tk.X, pady=4)
        ttk.Button(bf, text="+ 添加映射", command=self._add_mapping_row).pack(side=tk.LEFT)
        ttk.Button(bf, text="🔄 重新自动检测", command=self._auto_map_fields).pack(side=tk.LEFT, padx=8)
        self.detect_st = tk.StringVar()
        ttk.Label(bf, textvariable=self.detect_st, foreground="#81c784").pack(side=tk.LEFT, padx=12)
        self.dup_st = tk.StringVar()
        ttk.Label(bf, textvariable=self.dup_st, foreground="#ff6d00",
                  font=("TkDefaultFont", 10, "bold")).pack(side=tk.LEFT, padx=12)
        ttk.Button(bf, text="⚙ DP→城市", command=self._open_dp_editor, width=12).pack(side=tk.RIGHT, padx=4)

        # ── 操作按钮 ──
        af = ttk.Frame(main); af.pack(fill=tk.X, pady=(0, 6))
        self.start_btn = ttk.Button(af, text="🚀 开始匹配", command=self._start_matching, state="disabled")
        self.start_btn.pack()
        self.prog_var = tk.StringVar()
        ttk.Label(af, textvariable=self.prog_var, foreground="#90caf9").pack(pady=2)
        # 匹配完成后的结果栏（初始隐藏）
        self.result_bar = ttk.Frame(af)
        self.result_bar.pack(pady=4)
        self.result_bar.pack_forget()

        # ── 日志 ──
        lf = ttk.LabelFrame(main, text="📋 运行日志", padding=6)
        lf.pack(fill=tk.BOTH, expand=True)
        self.log_text = tk.Text(lf, height=10, font=("Consolas", 10), wrap=tk.WORD,
                                bg="#1e1e1e", fg="#d4d4d4", insertbackground="white",
                                relief=tk.FLAT, borderwidth=0)
        self.log_text.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        sb = ttk.Scrollbar(lf, command=self.log_text.yview)
        sb.pack(fill=tk.Y, side=tk.RIGHT)
        self.log_text.configure(yscrollcommand=sb.set)
        for tag, color in [("info", "#90caf9"), ("good", "#81c784"), ("warn", "#ff6d00"), ("err", "#ef5350")]:
            self.log_text.tag_configure(tag, foreground=color)

        self._update_sheet_checkboxes([])

    def _on_map_inner_configure(self, event):
        self.map_canvas.configure(scrollregion=self.map_canvas.bbox("all"))

    def _on_map_canvas_configure(self, event):
        # Keep inner frame width matched to canvas width
        self.map_canvas.itemconfig(self.map_canvas_window, width=event.width)

    # ==================== 日志 ====================
    def log(self, msg, tag="info"):
        self.log_text.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n", tag)
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    # ==================== 文件加载 ====================
    def _select_src(self):
        path = filedialog.askopenfilename(title="选择 Main HUB XLSX 文件",
                                          filetypes=[("Excel", "*.xlsx *.xls"), ("所有", "*.*")])
        if not path: return
        self.src_path_var.set(path); self.src_path = path
        try:
            self.src_wb = openpyxl.load_workbook(path)
            self.log(f"已加载源文件: {os.path.basename(path)}", "good")
            candidates = _find_candidate_sheets(self.src_wb)
            self.log(f"可用 Sheet: {', '.join(candidates)}", "info")
            self.selected_sheets = [s for s in ["LEJ", "LGG"] if s in candidates]
            if not self.selected_sheets and candidates:
                self.selected_sheets = [candidates[0]]
            self._update_sheet_checkboxes(candidates)
            self._load_source()
        except Exception as e:
            self.log(f"加载失败: {e}", "err")
            messagebox.showerror("错误", str(e))

    def _select_tgt(self):
        path = filedialog.askopenfilename(title="选择 DHL 用车记录 XLSX 文件",
                                          filetypes=[("Excel", "*.xlsx *.xls"), ("所有", "*.*")])
        if not path: return
        self.tgt_path_var.set(path); self.tgt_path = path
        try:
            self.tgt_wb = openpyxl.load_workbook(path)
            self.tgt_sheet = self.tgt_wb.sheetnames[0]
            self.tgt_headers, self.tgt_rows = read_sheet_rows(self.tgt_wb, self.tgt_sheet)
            self.log(f"已加载目标文件: {os.path.basename(path)} ({len(self.tgt_rows)} 行)", "good")
            self._refresh_ui()
        except Exception as e:
            self.log(f"加载失败: {e}", "err")
            messagebox.showerror("错误", str(e))

    def _update_sheet_checkboxes(self, sheets):
        for w in self.shcf.winfo_children(): w.destroy()
        self.sheet_vars = {}
        for s in sheets:
            v = tk.BooleanVar(value=s in self.selected_sheets)
            self.sheet_vars[s] = v
            ttk.Checkbutton(self.shcf, text=s, variable=v,
                            command=self._on_sheets_changed).pack(side=tk.LEFT, padx=4, pady=4)
        if not sheets:
            ttk.Label(self.shcf, text="请先加载源文件", foreground="gray").pack(side=tk.LEFT)

    def _on_sheets_changed(self):
        self.selected_sheets = [s for s, v in self.sheet_vars.items() if v.get()]
        if self.selected_sheets:
            self._load_source()
        else:
            self.src_st.set("未选择 Sheet")
            self._refresh_ui()

    def _load_source(self):
        if not self.src_wb or not self.selected_sheets: return
        try:
            self.src_headers, self.src_row_count, self.src_images = read_source_data(self.src_wb, self.selected_sheets)
            total = self.src_row_count
            self.src_st.set(f"{', '.join(self.selected_sheets)} · {total} 行")

            # ──── 重复检测 ────
            self._has_cross_sheet_dups = False
            self.dup_st.set("")
            dups = check_duplicates(self.selected_sheets, self.src_wb)
            if dups:
                self._has_cross_sheet_dups = True
                n = len(dups)
                self.dup_st.set(f"⚠ 跨 Sheet 重复: {n} 个 Reference")
                self.log(f"⚠ 检测到 {n} 个跨 Sheet 重复的 Loading Reference:", "warn")
                for ref, locs in list(dups.items())[:10]:
                    loc_str = " / ".join(f"{s}第{r}行" for s, r in locs)
                    self.log(f"  · {ref} → {loc_str}", "warn")
                if n > 10: self.log(f"  ... 还有 {n-10} 个", "warn")
                self.log("  请先在源文件中解决重复问题，再重新匹配。", "warn")
            else:
                self.log("✓ 无跨 Sheet 重复", "good")

            self.log(f"源数据: {total} 行, 图片位置: {len(self.src_images)}", "info")
            self._refresh_ui()
        except Exception as e:
            self.log(f"读取源数据失败: {e}", "err")

    # ==================== UI 刷新 ====================
    def _refresh_ui(self):
        ok = bool(self.src_headers and self.tgt_headers)
        self.start_btn.configure(state="normal" if ok else "disabled")
        if not ok: return

        src_opts = [norm(h) for h in self.src_headers if h and h != "None"]
        tgt_opts = [norm(h) for h in self.tgt_headers if h and h != "None"]
        self.src_key_cb["values"] = src_opts
        self.tgt_key_cb["values"] = tgt_opts

        ds = next((h for h in src_opts if "loading reference" in h.lower() or "提货码" in h.lower()),
                  src_opts[0] if src_opts else "")
        dt = next((h for h in tgt_opts if "dhl reference" in h.lower()), tgt_opts[0] if tgt_opts else "")
        self.src_key_var.set(ds); self.tgt_key_var.set(dt)
        self._auto_map_fields()

    def _auto_map_fields(self):
        for w in self.map_inner.winfo_children(): w.destroy()
        self.mapping_rows.clear()
        sk = self.src_key_var.get(); tk = self.tgt_key_var.get()
        found = 0
        for sh, th in DEFAULT_MAPPINGS:
            si = fuzzy_find(self.src_headers, sh)
            ti = fuzzy_find(self.tgt_headers, th)
            if si != -1 and ti != -1:
                sn = norm(self.src_headers[si]); tn = norm(self.tgt_headers[ti])
                if sn == sk or tn == tk: continue
                self._add_mapping_row(sn, tn)
                found += 1
        self.detect_st.set(f"自动检测到 {found} 对映射")

    def _add_mapping_row(self, sv="", tv=""):
        rf = ttk.Frame(self.map_inner); rf.pack(fill=tk.X, pady=1)
        idx = len(self.mapping_rows) + 1
        ttk.Label(rf, text=str(idx), width=3).pack(side=tk.LEFT)
        src_var = tk.StringVar(value=sv)
        src_cb = ttk.Combobox(rf, textvariable=src_var, width=36, state="readonly")
        src_cb["values"] = [norm(h) for h in self.src_headers if h and h != "None"] if self.src_headers else []
        src_cb.pack(side=tk.LEFT, padx=4)
        ttk.Label(rf, text="→", width=3).pack(side=tk.LEFT)
        tgt_var = tk.StringVar(value=tv)
        tgt_cb = ttk.Combobox(rf, textvariable=tgt_var, width=36, state="readonly")
        tgt_cb["values"] = [norm(h) for h in self.tgt_headers if h and h != "None"] if self.tgt_headers else []
        tgt_cb.pack(side=tk.LEFT, padx=4)

        def rm():
            if len(self.mapping_rows) <= 1: return
            rf.destroy(); self.mapping_rows.remove(rd); self._renumber()
        rd = {"frame": rf, "src_var": src_var, "tgt_var": tgt_var}
        ttk.Button(rf, text="✕", width=3, command=rm).pack(side=tk.LEFT)
        self.mapping_rows.append(rd)

        # 滚动到底部
        self.map_canvas.after(50, lambda: self.map_canvas.yview_moveto(1.0))

    def _renumber(self):
        for i, mr in enumerate(self.mapping_rows):
            ch = mr["frame"].winfo_children()
            if ch: ch[0].configure(text=str(i + 1))

    # ==================== 匹配 ====================
    def _open_dp_editor(self):
        """弹出窗口编辑 DP代码 → 城市的映射字典"""
        dlg = tk.Toplevel(self.root)
        dlg.title("DP 代码 → 目的地城市 映射")
        dlg.geometry("440x520")
        dlg.resizable(True, True)
        dlg.transient(self.root)
        dlg.grab_set()

        main = ttk.Frame(dlg, padding=10)
        main.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main, text="DP代码 → 城市名，匹配时从 Reference 自动解析 Destination",
                  foreground="gray").pack(anchor=tk.W, pady=(0, 8))

        # 表格区 — 限制高度，确保底部按钮始终可见
        table_frame = ttk.Frame(main)
        table_frame.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(table_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        self._dp_inner = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=self._dp_inner, anchor="nw")
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def _on_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        self._dp_inner.bind("<Configure>", _on_configure)

        # 按钮区 — 固定高度，始终可见
        bf = ttk.Frame(main)
        bf.pack(fill=tk.X, pady=(8, 0))

        ttk.Button(bf, text="+ 添加行", command=self._dp_add_row).pack(side=tk.LEFT)
        ttk.Button(bf, text="💾 保存", command=lambda: self._dp_save(dlg)).pack(side=tk.RIGHT)
        ttk.Button(bf, text="取消", command=dlg.destroy).pack(side=tk.RIGHT, padx=4)

        self._dp_entries = []
        self._dp_canvas = canvas
        self._dp_populate()

    def _dp_populate(self):
        for w in self._dp_inner.winfo_children():
            w.destroy()
        self._dp_entries.clear()
        # 表头
        ttk.Label(self._dp_inner, text="DP 代码", width=12,
                  font=("TkDefaultFont", 9, "bold")).grid(row=0, column=0, padx=2, pady=2)
        ttk.Label(self._dp_inner, text="→", width=3).grid(row=0, column=1)
        ttk.Label(self._dp_inner, text="目的地城市", width=22,
                  font=("TkDefaultFont", 9, "bold")).grid(row=0, column=2, padx=2, pady=2)
        for i, (code, city) in enumerate(sorted(self.dp_map.items())):
            self._dp_add_entry_row(i + 1, code, city)

    def _dp_add_entry_row(self, row_idx, code="", city=""):
        cv = tk.StringVar(value=code)
        nv = tk.StringVar(value=city)
        ttk.Entry(self._dp_inner, textvariable=cv, width=14).grid(row=row_idx, column=0, padx=2, pady=1)
        ttk.Label(self._dp_inner, text="→", width=3).grid(row=row_idx, column=1)
        ttk.Entry(self._dp_inner, textvariable=nv, width=24).grid(row=row_idx, column=2, padx=2, pady=1)
        self._dp_entries.append((cv, nv))
        self._dp_canvas.yview_moveto(1.0)

    def _dp_add_row(self):
        row_idx = len(self._dp_entries) + 1
        self._dp_add_entry_row(row_idx)
        # 强制更新 scrollregion
        self._dp_canvas.configure(scrollregion=self._dp_canvas.bbox("all"))

    def _dp_save(self, dlg):
        new_map = {}
        for cv, nv in self._dp_entries:
            c = cv.get().strip()
            n = nv.get().strip()
            if c and n:
                new_map[c] = n
        if new_map:
            self.dp_map = new_map
            save_dp_map(self.dp_map)
            self.log(f"DP 映射已更新: {len(self.dp_map)} 条", "info")
        dlg.destroy()

    def _start_matching(self):
        if not self.mapping_rows:
            messagebox.showwarning("提示", "请至少配置一对字段映射")
            return

        # 如果有跨Sheet重复，直接拒绝
        if self._has_cross_sheet_dups:
            messagebox.showwarning("无法匹配", f"{self.dup_st.get()}\n\n请先解决跨 Sheet 重复问题，再开始匹配。")
            return

        self.start_btn.configure(state="disabled", text="匹配中...")
        self.prog_var.set("正在处理...")
        self.result_bar.pack_forget()
        threading.Thread(target=self._do_match, daemon=True).start()

    def _do_match(self):
        try:
            # 复用 _select_src 阶段已打开的源文件，不再重复加载
            _src_wb = self.src_wb
            _headers, _, _images = read_source_data(_src_wb, self.selected_sheets)

            # 输出文件：先复制目标文件，保留其中已有的图片（邮件截图等）
            out = os.path.splitext(self.tgt_path)[0] + "_matched.xlsx"
            shutil.copy2(self.tgt_path, out)

            # 确保输出文件可写（源文件可能是只读的，copy2 会继承权限）
            os.chmod(out, 0o666)

            # 打开复制后的文件进行修改

            skn = self.src_key_var.get(); tkn = self.tgt_key_var.get()
            ski = next((i for i, h in enumerate(_headers) if norm(h) == skn), -1)
            tki = next((i for i, h in enumerate(self.tgt_headers) if norm(h) == tkn), -1)
            if ski == -1 or tki == -1: self._uilog("❌ 匹配键未找到", "err"); self._uidone(); return

            self._uilog(f"源键: {skn} (列{ski})  |  目标键: {tkn} (列{tki})", "info")

            # 解析映射
            cmaps = []
            for mr in self.mapping_rows:
                si = next((i for i, h in enumerate(_headers) if norm(h) == mr["src_var"].get()), -1)
                ti = next((i for i, h in enumerate(self.tgt_headers) if norm(h) == mr["tgt_var"].get()), -1)
                if si != -1 and ti != -1: cmaps.append((si, ti, mr["src_var"].get(), mr["tgt_var"].get()))
            self._uilog(f"字段映射: {len(cmaps)} 对", "info")
            for si, ti, sn, tn in cmaps: self._uilog(f"  {sn} → {tn}", "info")

            # 源索引 — 流式遍历，跳过删除线行；遇到真实重复直接终止
            idx = {}
            for sn_ in self.selected_sheets:
                ws = _src_wb[sn_]
                for ri_, row in enumerate(ws.iter_rows(min_row=2)):
                    ref_cell = row[ski] if ski < len(row) else None
                    if ref_cell is None or ref_cell.value is None: continue
                    if is_strikethrough(ref_cell): continue
                    k = str(ref_cell.value).strip()
                    if k and is_valid_reference(k):
                        if k in idx:
                            prev_sn, prev_ri = idx[k][1], idx[k][2]
                            self._uilog(
                                f"❌ 重复 Reference: {k}\n"
                                f"    位置1: {prev_sn} 第{prev_ri + 2}行\n"
                                f"    位置2: {sn_} 第{ri_ + 2}行\n"
                                f"  请解决重复后再匹配。", "err")
                            self._uidone()
                            self.root.after(0, lambda: messagebox.showerror(
                                "匹配失败",
                                f"源数据中存在重复的 Loading Reference:\n\n{k}\n\n"
                                f"位置1: {prev_sn} 第{prev_ri + 2}行\n"
                                f"位置2: {sn_} 第{ri_ + 2}行\n\n"
                                f"请先在源文件中解决重复问题，再重新匹配。"))
                            return
                        row_vals = [c.value if c.value is not None else "" for c in row]
                        idx[k] = (row_vals, sn_, ri_)
            self._uilog(f"源索引: {len(idx)} 条唯一 Key", "info")

            # 匹配
            matched = unmatched = skipped = 0

            # 重新加载目标数据与写入目标对齐
            self.tgt_wb = openpyxl.load_workbook(out)
            tws = self.tgt_wb[self.tgt_sheet]
            _, _tgt_rows = read_sheet_rows(self.tgt_wb, self.tgt_sheet)

            note_si = next((i for i, h in enumerate(_headers) if norm(h).lower() == "note"), -1)
            note_ti = next((i for i, h in enumerate(self.tgt_headers) if norm(h).lower() == "note"), -1)
            note_mapped = any(si == note_si and ti == note_ti for si, ti, _, _ in cmaps)

            # Destination 字段不从源文件取值，改为从 Reference 解析
            dest_ti = next((ti for si, ti, sn, tn in cmaps if "destination" in tn.lower()), -1)

            for ti_, row in enumerate(_tgt_rows):
                k = str(row[tki]).strip() if tki < len(row) else ""
                if not k: skipped += 1; continue
                if k not in idx: unmatched += 1; continue
                srow, shn, sri = idx[k]
                matched += 1
                for sc, tc, sn, tn in cmaps:
                    if tc == dest_ti:
                        v = extract_destination(k, self.dp_map)  # 从 Reference 解析目的地
                    else:
                        v = srow[sc] if sc < len(srow) else ""
                    if v is not None and str(v).strip():
                        row[tc] = v
                        tws.cell(row=ti_ + 2, column=tc + 1).value = v

                if note_mapped and note_si != -1 and note_ti != -1:
                    copy_images_for_row(_src_wb, shn, sri, tws, ti_ + 2,
                                       note_si, note_ti, _images)

            self._uilog("", ""); self._uilog("=" * 40, "info")
            self._uilog(f"  ✅ 匹配成功: {matched} 行", "good")
            if unmatched: self._uilog(f"  ⚠ 未匹配: {unmatched} 行", "warn")
            if skipped: self._uilog(f"  ⊝ 跳过空键: {skipped} 行", "info")
            self._uilog("=" * 40, "info")

            self.tgt_wb.save(out)
            self._uilog(f"📥 输出: {os.path.basename(out)}", "good")
            self._output_path = out
            self._uidone()
            # 显示结果栏
            self.root.after(0, lambda: self._show_result_bar(out))
            self.root.after(0, lambda: messagebox.showinfo(
                "完成",
                f"✅ 成功: {matched} 行  |  ⚠ 未匹配: {unmatched} 行  |  ⊝ 跳过: {skipped} 行\n"
                f"\n输出: {os.path.basename(out)}"))

        except Exception as e:
            import traceback
            self._uilog(f"❌ 错误: {e}", "err")
            self._uilog(traceback.format_exc()[-600:], "err")
            self._uidone()

    # ==================== 结果栏 ====================

    def _show_result_bar(self, output_path):
        """匹配完成后显示结果操作按钮"""
        for w in self.result_bar.winfo_children(): w.destroy()
        ttk.Label(self.result_bar, text="📥 结果就绪：",
                  font=("TkDefaultFont", 10, "bold"), foreground="#81c784").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(self.result_bar, text="📂 打开文件位置",
                   command=lambda: self._open_folder(output_path)).pack(side=tk.LEFT, padx=4)
        ttk.Button(self.result_bar, text="📄 打开匹配文件",
                   command=lambda: self._open_file(output_path)).pack(side=tk.LEFT, padx=4)
        self.result_bar.pack(pady=4)

    @staticmethod
    def _open_file(path):
        import subprocess, sys
        if sys.platform == "darwin":
            subprocess.Popen(["open", path])
        elif sys.platform == "win32":
            os.startfile(path)
        else:
            subprocess.Popen(["xdg-open", path])

    @staticmethod
    def _open_folder(path):
        import subprocess, sys
        folder = os.path.dirname(os.path.abspath(path))
        if sys.platform == "darwin":
            subprocess.Popen(["open", folder])
        elif sys.platform == "win32":
            os.startfile(folder)
        else:
            subprocess.Popen(["xdg-open", folder])

    def _uilog(self, m, t="info"): self.root.after(0, lambda: self.log(m, t))

    def _uidone(self):
        self.root.after(0, lambda: self.start_btn.configure(state="normal", text="🚀 开始匹配"))
        self.root.after(0, lambda: self.prog_var.set(""))

# ============================================================ 入口 ============================================================

if __name__ == "__main__":
    root = tk.Tk()
    MatchApp(root)
    root.mainloop()
